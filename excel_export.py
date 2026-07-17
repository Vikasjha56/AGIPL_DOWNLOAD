import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment
)

from openpyxl.utils import get_column_letter

from datetime import datetime
import pytz


# =====================================================
# AGIPL THEME
# =====================================================

HEADER_COLOR = "0B3B6F"
WHITE = "FFFFFF"
BORDER_COLOR = "C9C9C9"


# =====================================================
# BORDER
# =====================================================

thin = Side(
    border_style="thin",
    color=BORDER_COLOR
)

border = Border(
    left=thin,
    right=thin,
    top=thin,
    bottom=thin
)


# =====================================================
# FONT
# =====================================================

title_font = Font(
    size=18,
    bold=True,
    color=WHITE
)

header_font = Font(
    size=11,
    bold=True,
    color=WHITE
)


# =====================================================
# CREATE EXCEL
# =====================================================

def create_excel(master_df):

    master_df = master_df.copy()


    # ==========================================
    # CLEAN COLUMN NAMES
    # ==========================================

    master_df.columns = (
        master_df.columns
        .astype(str)
        .str.strip()
    )


    master_df = master_df.dropna(
        how="all"
    )


    # ==========================================
    # OWNED DATA ONLY
    # ==========================================

    if "Owned/Hired" in master_df.columns:

        master_df = master_df[
            master_df["Owned/Hired"]
            .astype(str)
            .str.strip()
            .str.upper()
            == "OWNED"
        ].copy()


    # ==========================================
    # RESET DATA INDEX
    # ==========================================

    master_df.reset_index(
        drop=True,
        inplace=True
    )


    # ==========================================
    # CREATE FRESH SERIAL INDEX NUMBER
    # ==========================================

    if "Index Number" in master_df.columns:

        master_df.drop(
            columns=["Index Number"],
            inplace=True
        )


    if "No" in master_df.columns:

        master_df.drop(
            columns=["No"],
            inplace=True
        )


    master_df.insert(
        0,
        "Index Number",
        range(
            1,
            len(master_df) + 1
        )
    )


    # ==========================================
    # WORKBOOK
    # ==========================================

    wb = Workbook()

    ws = wb.active

    ws.title = "Breakdown Report"


    last_column = get_column_letter(
        len(master_df.columns)
    )


    ws.merge_cells(
        f"A1:{last_column}1"
    )


    title = ws["A1"]

    title.value = (
        "AGIPL BREAKDOWN PENDING REPORT"
    )

    title.font = title_font

    title.fill = PatternFill(
        fill_type="solid",
        fgColor=HEADER_COLOR
    )

    title.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )


    ws.row_dimensions[1].height = 30



    # ==========================================
    # REPORT DATE
    # ==========================================

    india_time = datetime.now(
        pytz.timezone(
            "Asia/Kolkata"
        )
    )


    ws["A2"] = "Report Generated"

    ws["B2"] = india_time.strftime(
        "%d-%m-%Y %I:%M:%S %p"
    )


    ws["A2"].font = Font(
        bold=True
    )



    # ==========================================
    # HEADER
    # ==========================================

    start_row = 5


    for col_num, column in enumerate(
        master_df.columns,
        start=1
    ):


        cell = ws.cell(
            row=start_row,
            column=col_num
        )


        cell.value = column

        cell.font = header_font

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor=HEADER_COLOR
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

        cell.border = border



    # ==========================================
    # DATA INSERT
    # ==========================================

    for row_num, row in enumerate(
        master_df.values,
        start=start_row + 1
    ):


        for col_num, value in enumerate(
            row,
            start=1
        ):


            cell = ws.cell(
                row=row_num,
                column=col_num
            )


            cell.value = value

            cell.border = border

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )



    # ==========================================
    # COLUMN WIDTH
    # ==========================================

    width_map = {


        "Index Number":12,

        "Site":20,

        "Date of breakdown":18,

        "Category":22,

        "Vehcile No":18,

        "Breakdown Details":45,

        "Reason for pendency":40,

        "Pending for (no of days)":18,

        "Owned/Hired":15,

        "Breakdown Alert Icon":20

    }



    for column_cells in ws.columns:


        column_number = (
            column_cells[0].column
        )


        column_letter = get_column_letter(
            column_number
        )


        header = ws.cell(
            row=start_row,
            column=column_number
        ).value



        if header in width_map:


            ws.column_dimensions[
                column_letter
            ].width = width_map[header]


        else:


            max_length = len(
                str(header)
            )


            for cell in column_cells:


                if cell.value is not None:

                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )


            ws.column_dimensions[
                column_letter
            ].width = min(
                max_length + 3,
                30
            )



    # ==========================================
    # FREEZE
    # ==========================================

    ws.freeze_panes = "A6"



    # ==========================================
    # FILTER
    # ==========================================

    last_column = get_column_letter(
        ws.max_column
    )


    ws.auto_filter.ref = (
        f"A5:{last_column}{ws.max_row}"
    )



    # ==========================================
    # ROW HEIGHT
    # ==========================================

    for row in range(
        6,
        ws.max_row + 1
    ):

        ws.row_dimensions[row].height = 40



    # ==========================================
    # SAVE FILE
    # ==========================================

    output_file = (
        "Pending_Breakdown_Report.xlsx"
    )


    wb.save(
        output_file
    )


    print(
        "Pending Excel Report Created Successfully"
    )


    return output_file
